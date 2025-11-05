import json
from itertools import islice

import openpyxl


def process_keywords_to_file(sources, output):
    with open(output, "w") as outfile:
        json.dump(process_keywords(sources), outfile, indent=4)


def process_keywords(sources):
    return merge_dictionaries(
        {source.get("key", "unknown"): process_source(source) for source in sources}
    )


def process_source(source):
    input_file = source.get("location") or source["path"]
    book = openpyxl.load_workbook(input_file, read_only=True)
    all_tables = {}

    for sheet in book.worksheets:
        try:
            all_tables[sheet.title] = process_sheet(sheet)
        except TableNotFoundError:
            continue

    return all_tables


def process_sheet(sheet):
    row_start = find_table_row_start(sheet) + 1
    misspellings_start = find_misspellings_col_start(sheet)

    return [
        wordset
        for row in sheet.iter_rows(min_row=row_start)
        if (
            wordset := create_wordset(misspellings_start, row)
        )
    ]


def create_wordset(misspellings_start, row):
    words = [
        read_cols(row, 1, misspellings_start),
        read_cols(row, misspellings_start),
    ]
    language = read_cols(row, 0, 1)
    if any(words):
        return {
            language[0]: {
                "keywords": words[0],
                "mispellings": words[1],
            },
        }
    else:
        return None


def find_table_row_start(sheet):
    """We assume that the table starts in the first column,
    in the left-most cell in the row before it we have a
    cell containing the matching string.

    Note: We assume there's only one such table per sheet.
    """
    MATCHING_STRING = (
        "Please insert translation of each word under each corresponding cell. If the "
        "particular word does not translate into the chosen language, please leave it "
        "blank"
    )
    index = index_of(
        [row[0] for row in sheet["A1":"A10"]],
        MATCHING_STRING,
    )

    if index == -1:
        raise TableNotFoundError()

    # to reference the next row
    return index + 1


def find_misspellings_col_start(sheet):
    HEADER2 = "Range of possible misspellings and common slang used by the population"
    index = -1

    for row in sheet.iter_rows():
        index = index_of(row, HEADER2)
        if index > -1:
            return index


def index_of(seq, value):
    try:
        return [i.value for i in seq].index(value)
    except ValueError:
        return -1


def read_cols(row, start, end=None):
    return [
        str(cell.value).strip() for cell in row[start:end] if cell.value is not None
    ]


def merge_dictionaries(dictionaries):
    """
    Merge multiple language dictionaries into a combined structure.
    """
    it = iter(dictionaries.items())
    first_lang, merged = next(it)

    # merged is a dict of sheets → list of rows
    for lang, dic in it:
        for sheet, rows in merged.items():
            if sheet not in dic:
                print(f"Sheet '{sheet}' not found in {lang}, skipping")
                continue

            target_rows = dic[sheet]

            # Assume rows are aligned by index (order in the sheet)
            if len(rows) != len(target_rows):
                print(f"Row count mismatch in sheet '{sheet}' between {first_lang} and {lang}")
                continue

            for i, row in enumerate(rows):
                # Add new language data directly as a top-level key
                merged[sheet][i][lang] = target_rows[i].get(lang, target_rows[i])

    return merged


def batch(iterable, n):
    """Batch data into tuples of length n.
    Stops when a batch has fewer than n items.
    batch('ABCDEFG', 3) --> ABC DEF"""
    if n < 1:
        raise ValueError("n must be at least one")
    it = iter(iterable)
    while len(batch := tuple(islice(it, n))) == n:
        yield batch


class TableNotFoundError(Exception):
    pass
